import numpy as np
from sklearn import preprocessing, cross_validation
from sklearn.ensemble import RandomForestClassifier
import pandas as pd
from dietplan import zone, fasting, keto, hc, hp, atkins
# import zone
# import fasting
# import keto
# import hc
# import hp
# import atkins
import openpyxl


#converting categorical data to numeric data
def handle_non_numerical_data(df):
    print("***********************Handling non-numerical data")
    columns = df.columns.values
    for column in columns:
        text_digit_vals = {}
        def convert_to_int(val):
            return text_digit_vals[val]

        if df[column].dtype != np.int64 and df[column].dtype != np.float64:
            column_contents = df[column].values.tolist()
            unique_elements = set(column_contents)
            x = 0
            for unique in unique_elements:
                if unique not in text_digit_vals:
                    text_digit_vals[unique] = x
                    x+=1

            df[column] = list(map(convert_to_int, df[column]))
    return df


def classify(user_calorieintake):
    print("************In rf.py")
    np.random.seed(0)
    #data frame for train and test data
    print("**********************reading Excel files.")
    df = pd.read_excel('train.xlsx')
    dft = pd.read_excel('test.xlsx')
    #print(df.head())

    df = handle_non_numerical_data(df)
    dft = handle_non_numerical_data(dft)
    #print(df.head())
    #print(dft.head())
    train = df
    test = dft

    #selecting required features
    features = df.columns[:6]
    featurest = dft.columns[:6]

    #applying the classifier
    print("**********************Classifying using random forest.")
    clf = RandomForestClassifier(n_jobs=2, random_state=0)
    clf.fit(train[features], train['Diet Plan'])
    print("**********************Predicting the dietplan")
    preds = clf.predict(test[featurest])

    testxl = openpyxl.load_workbook("test.xlsx")

    test_sheet = testxl.get_sheet_by_name("Sheet1")

    #print(preds[-1])
    #appending the predicted result to the test sheet and generating the corresponding diet plan
    if preds[-1] == 0:
        print("************Intermittent Fasting")
        test_sheet.cell(row = test_sheet.max_row, column = 8).value = "Intermittent Fasting"
        meals = fasting.main(user_calorieintake)
    elif preds[-1] == 1:
        print("************High Protein - Low Carb Diet")
        test_sheet.cell(row = test_sheet.max_row, column = 8).value = 'High Protein - Low Carb Diet'
        meals = hp.main(user_calorieintake)
    elif preds[-1] == 2:
        print("************Zone Diet")
        test_sheet.cell(row = test_sheet.max_row, column = 8).value = "Zone Diet"
        meals = zone.main(user_calorieintake)
    elif preds[-1] == 3:
        print("************Atkins Diet")
        test_sheet.cell(row = test_sheet.max_row, column = 8).value = "Atkins Diet"
        meals = atkins.main(user_calorieintake)
    elif preds[-1] == 4:
        print("************High Carbs - Low Fat Diet")
        test_sheet.cell(row = test_sheet.max_row, column = 8).value = "High Carbs - Low Fat Diet"
        meals = hc.main(user_calorieintake)
    else:
        print("************Ketogenic Diet")
        test_sheet.cell(row = test_sheet.max_row, column = 8).value = "Ketogenic Diet"
        meals = keto.main(user_calorieintake)
    testxl.save("test.xlsx")
    print("************************Generated plan************************")
    print(meals)
    print("**************************************************************")
    return meals

if __name__ == '__main__':
    classify()
