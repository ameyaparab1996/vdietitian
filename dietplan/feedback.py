import openpyxl

#copying the row from test sheet
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow,endRow + 1,1):
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected

#pasting the feedback to the last row of test sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def handlefeedback(feedback_is, cust_ID):
    test = openpyxl.load_workbook("test.xlsx")
    print("***********test done")
    test_sheet = test.get_sheet_by_name("Sheet1")
    print("***********test_sheet done")
    train = openpyxl.load_workbook("train.xlsx")
    train_sheet = train.get_sheet_by_name("Form Responses 1")
    print("***********train_sheet done")

    i = 2
    #print("***********i is 2")
    while i <= test_sheet.max_row:
        #print("***********in while")
        if test_sheet.cell(row = i, column = 9).value == cust_ID:
            #print("***********in while in if")
            feedback_row = i
            #print("***********feedback_row is i")
        i = i+1
        #print("***********i is i+1")
    #adding the positive feedback to the training data
    #print("feedback_is: "+feedback_is)
    if feedback_is == 'Positive':
        selectedRange = copyRange(1,feedback_row,8,feedback_row,test_sheet)
        pastingRange = pasteRange(1,train_sheet.max_row+1,8,train_sheet.max_row+1,train_sheet,selectedRange)
        train.save("train.xlsx")
        print("************Positive feedback added to train data")
    elif feedback_is == 'Negative':
        print("************Feedback is Negative")
        age = test_sheet.cell(row = feedback_row, column = 1).value
        gender = test_sheet.cell(row = feedback_row, column = 2).value
        height = test_sheet.cell(row = feedback_row, column = 3).value
        weight = test_sheet.cell(row = feedback_row, column = 4).value
        goal = test_sheet.cell(row = feedback_row, column = 5).value
        lifestyle = test_sheet.cell(row = feedback_row, column = 6).value
        diet = test_sheet.cell(row = feedback_row, column = 8).value
        i = 2
        while i <= train_sheet.max_row:
            age_t = train_sheet.cell(row = i, column = 1).value
            gender_t = train_sheet.cell(row = i, column = 2).value
            height_t = train_sheet.cell(row = i, column = 3).value
            weight_t = train_sheet.cell(row = i, column = 4).value
            goal_t = train_sheet.cell(row = i, column = 5).value
            lifestyle_t = train_sheet.cell(row = i, column = 6).value
            diet_t = train_sheet.cell(row = i, column = 8).value
            if age_t >= age-5 and age_t <= age+5 and gender_t == gender and height_t >= height-10 and height <= height+10 and weight_t >= weight-5 and weight_t <= weight+5 and goal_t == goal and lifestyle_t == lifestyle and diet_t == diet:
    			#delete row
                train_sheet.delete_rows(i, 1)
            i = i+1
        train.save("train.xlsx")
        print("************Negative data deleted from train data")
